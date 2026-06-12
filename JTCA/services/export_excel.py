"""
============================================================
JTCA - Services: Excel Export
Generates SAP_Export.xlsx with all shipment data
============================================================
"""

import logging
import os
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

_BASE_DIR = Path(__file__).parent.parent
EXPORT_DIR = _BASE_DIR / "data" / "exports"


def export_to_excel(
    shipments: list[dict],
    output_path: str | None = None,
) -> str:
    """
    Export shipments to a SAP-compatible Excel file.

    Args:
        shipments: List of shipment dicts from database
        output_path: Optional custom output path; defaults to data/exports/

    Returns:
        Absolute path of the created Excel file
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        logger.error(f"Missing export dependency: {e}")
        raise

    if not shipments:
        logger.warning("No shipments to export.")
        raise ValueError("No shipments available to export.")

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(EXPORT_DIR / f"SAP_Export_{timestamp}.xlsx")

    # ── Build DataFrame ──────────────────────────────────────
    rows = []
    for s in shipments:
        hscode = s.get("suggested_hs_code", "")
        target_sap = s.get("target_sap_system", "")
        if not target_sap:
            target_sap = "Table_ZLANDED_COST" if hscode.startswith("9025") else "Condition_Type_ZDUT"

        rows.append({
            "Shipment_ID": s.get("shipment_id", ""),
            "Manufacturing_Part_Number": s.get("part_number", ""),
            "Product_Description": s.get("product_description", ""),
            "Material_Type": s.get("material_type", "ZROH"),
            "Plant_Code": s.get("plant_code", "US02"),
            "Supplier_Name": s.get("supplier_name", "EMERSON"),
            "Country_of_Origin": s.get("country_of_origin", ""),
            "Shipping_Country": s.get("shipping_country", "Malaysia"),
            "WTO_Member_Status": s.get("wto_member_status", "Yes"),
            "FTA_Applicable": s.get("fta_applicable", "No"),
            "HTS_Code": hscode,
            "Declared_Value_USD": s.get("declared_value", 0.0),
            "Tariff_Rate_Percent": s.get("tariff_percent", 0.0),
            "Estimated_Duty_USD": s.get("estimated_duty", 0.0),
            "AI_Confidence": f"{s.get('confidence_score', 0.0):.0f}%",
            "Target_SAP_System": target_sap,
            "Status": s.get("status", ""),
        })

    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False, sheet_name="JTCA_Export")

    # ── Style the workbook ───────────────────────────────────
    wb = load_workbook(output_path)
    ws = wb.active

    # Header styling — Jabil Blue
    header_fill = PatternFill(
        start_color="0057A8", end_color="0057A8", fill_type="solid"
    )
    header_font = Font(
        name="Calibri", bold=True, color="FFFFFF", size=11
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Apply header styles
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        # Auto-width
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # Data rows
    alt_fill = PatternFill(
        start_color="EBF4FF", end_color="EBF4FF", fill_type="solid"
    )
    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")

    status_colors = {
        "Approved": "D1FAE5",
        "Pending Review": "FEF3C7",
        "Rejected": "FEE2E2",
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, cell in enumerate(row, start=1):
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_alignment

            # Color Status column
            col_name = df.columns[col_idx - 1] if col_idx <= len(df.columns) else ""
            if col_name == "Status" and cell.value in status_colors:
                cell.fill = PatternFill(
                    start_color=status_colors[cell.value],
                    end_color=status_colors[cell.value],
                    fill_type="solid",
                )
            elif is_alt:
                cell.fill = alt_fill

    # Freeze header row
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35

    # ── Summary Sheet ────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.title = "Summary"

    total = len(shipments)
    approved = sum(1 for s in shipments if s.get("status") == "Approved")
    pending = sum(1 for s in shipments if s.get("status") == "Pending Review")
    rejected = sum(1 for s in shipments if s.get("status") == "Rejected")
    total_duties = sum(s.get("estimated_duty", 0.0) for s in shipments)
    avg_confidence = (
        sum(s.get("confidence_score", 0.0) for s in shipments) / total
        if total > 0 else 0
    )

    summary_data = [
        ["JTCA Export Summary", ""],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Total Shipments", total],
        ["Approved", approved],
        ["Pending Review", pending],
        ["Rejected", rejected],
        ["Total Estimated Duties (USD)", f"${total_duties:,.2f}"],
        ["Average Confidence Score", f"{avg_confidence:.1f}%"],
    ]

    for row in summary_data:
        ws_summary.append(row)

    # Style summary
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=14, color="0057A8")
    for row in ws_summary.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Calibri", size=11)
    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 25

    wb.save(output_path)
    logger.info(f"Excel exported: {output_path} ({len(shipments)} shipments)")
    return output_path
